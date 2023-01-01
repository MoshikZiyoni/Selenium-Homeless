from selenium import webdriver
import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import requests
import io
import pandas as pd
import urllib.request

service_obj=Service("C:/Users/moshi/Downloads/chromedriver.exe")

driver=webdriver.Chrome(service=service_obj)

driver.get('https://www.homeless.co.il/sale/')
driver.maximize_window()
time.sleep(5)
ccc=0
prices = []
properties = []
cities = []
rooms = []
floors = []
neighbourhoods = []
streets = []
names = []
for i in range (5):
    ccc=int(ccc)+2
    i=i+1
    print (i)
    # print (ccc,'Hello')
    if i==32:
        driver.find_element(By.CSS_SELECTOR,'a.pagingtext').click()
        ccc=0
        time.sleep(4)
        continue
    else:
        ccc=str(ccc)
        i=str(i)
        name=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[3]').text
        property=(driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[3]').text)
        city=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[4]').text
        neighbourhood=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[5]').text
        room=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[7]').text
        floor=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[8]').text
        price=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[9]').text
        street=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[10]/div/div[1]/div[2]/table/tbody/tr['+ccc+']/td[6]').text
        image=driver.find_element(By.XPATH,'/html/body/form/div[4]/div[4]/div/div[8]/div/a['+i+']/img')
        image_url = image.get_attribute("src")
        response = requests.get(image_url)

        with open(i+"image.jpg", "wb") as f:
            f.write(response.content)


        time.sleep(0.5)
        streets.append(street)
        rooms.append(float(room))
        properties.append(property)
        cities.append(city)
        neighbourhoods.append(neighbourhood)
        if floor =='קרקע':
            floor=int(0)
        else:
            floor=int(floor)
        floors.append(floor)
        if price=='':
            price=0
            prices.append(price)
        else:
            price = price.replace("₪", "")
            price = int(price.replace(",", ""))
            prices.append(price)


import openpyxl
# # Create a new Excel file
# workbook = openpyxl.Workbook()

# # Add a worksheet to the file
# worksheet = workbook.active
# worksheet.title = "Sheet1"

# # Save the file
# workbook.save("moshik1234.xlsx")


####START#########
# Create a new Excel file
workbook = openpyxl.load_workbook('moshik1234.xlsx')

# Access a worksheet
worksheet = workbook['Sheet1']

# Set the default font
font = openpyxl.styles.Font(name='Arial')
worksheet.font = font
row = 1
for name in properties:
    worksheet.cell(row=row, column=1).value = name
    row += 1
row=1
for name in cities:
    worksheet.cell(row=row, column=2).value = name
    row += 1
row=1
for name in neighbourhoods:
    worksheet.cell(row=row, column=3).value = name
    row += 1
row=1
for name in streets:
    worksheet.cell(row=row, column=4).value = name
    row += 1
row=1
for name in rooms:
    worksheet.cell(row=row, column=5).value = name
    row += 1
row=1
for name in floors:
    worksheet.cell(row=row, column=6).value = name
    row += 1
row=1
for name in names:
    worksheet.cell(row=row, column=7).value = name
    row += 1  
      
# Save the file
workbook.save('moshik1234.xlsx')

# Save and close the Excel file
workbook.close()


